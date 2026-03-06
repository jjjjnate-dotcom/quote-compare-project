from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader, simpleSplit
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas

FONT_KR = "HYSMyeongJo-Medium"
DEFAULT_COL_WIDTH = 8.43
DEFAULT_ROW_HEIGHT = 15.0


def _register_fonts() -> None:
    pdfmetrics.registerFont(UnicodeCIDFont(FONT_KR))


def _column_width_points(ws, col_idx: int) -> float:
    width = ws.column_dimensions[get_column_letter(col_idx)].width
    if width is None:
        width = DEFAULT_COL_WIDTH
    return width * 7.0 + 5.0


def _row_height_points(ws, row_idx: int) -> float:
    height = ws.row_dimensions[row_idx].height
    if height is None:
        height = DEFAULT_ROW_HEIGHT
    return float(height)


def _render_start_col(ws) -> int:
    # The wide "해광" template starts at column C.
    if ws.max_column >= 30:
        return 3
    return 1


def _used_bounds(ws) -> tuple[int, int, int, int]:
    min_row = 1
    min_col = _render_start_col(ws)
    max_row = 1
    max_col = min_col

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value not in (None, ""):
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)

    for image in getattr(ws, "_images", []):
        anchor = image.anchor
        max_row = max(max_row, anchor.to.row + 1)
        max_col = max(max_col, anchor.to.col + 1)

    return min_row, min_col, max_row, max_col


def _positions(
    ws, min_row: int, min_col: int, max_row: int, max_col: int
) -> tuple[dict[int, float], dict[int, float], float, float]:
    x_positions: dict[int, float] = {min_col - 1: 0.0}
    for col_idx in range(min_col, max_col + 1):
        x_positions[col_idx] = x_positions[col_idx - 1] + _column_width_points(ws, col_idx)

    y_positions: dict[int, float] = {min_row - 1: 0.0}
    for row_idx in range(min_row, max_row + 1):
        y_positions[row_idx] = y_positions[row_idx - 1] + _row_height_points(ws, row_idx)

    return x_positions, y_positions, x_positions[max_col], y_positions[max_row]


def _fill_color(fill: PatternFill):
    if fill is None or fill.patternType not in {"solid", "gray125"}:
        return None
    rgb = getattr(fill.fgColor, "rgb", None)
    if not rgb:
        return None
    rgb = str(rgb)
    if len(rgb) == 8:
        rgb = rgb[2:]
    if len(rgb) != 6:
        return None
    try:
        r = int(rgb[0:2], 16) / 255
        g = int(rgb[2:4], 16) / 255
        b = int(rgb[4:6], 16) / 255
    except ValueError:
        return None
    return colors.Color(r, g, b)


def _border_width(style_name: str | None) -> float:
    mapping = {
        "thin": 0.4,
        "medium": 0.8,
        "thick": 1.2,
        "double": 1.0,
        "hair": 0.2,
        "dashed": 0.4,
        "dotted": 0.4,
    }
    return mapping.get(style_name or "", 0.0)


def _draw_border_line(c: canvas.Canvas, x1: float, y1: float, x2: float, y2: float, width: float) -> None:
    if width <= 0:
        return
    c.setStrokeColor(colors.black)
    c.setLineWidth(width)
    c.line(x1, y1, x2, y2)


def _draw_text(c: canvas.Canvas, cell, x: float, y: float, width: float, height: float) -> None:
    value = cell.value
    if value in (None, ""):
        return

    if isinstance(value, datetime):
        text = value.strftime("%Y-%m-%d")
    elif isinstance(value, date):
        text = value.strftime("%Y-%m-%d")
    elif isinstance(value, (int, float)):
        if float(value).is_integer():
            if abs(float(value)) >= 1000 or "," in (cell.number_format or ""):
                text = f"{int(value):,}"
            else:
                text = str(int(value))
        else:
            text = f"{value:,.2f}".rstrip("0").rstrip(".")
    else:
        text = str(value)

    font_size = float(getattr(cell.font, "sz", 10) or 10)
    font_size = max(6.0, min(font_size, 11.0))
    padding_x = 2.0
    padding_y = 1.5

    lines = simpleSplit(text, FONT_KR, font_size, max(width - padding_x * 2, 8))
    if not lines:
        return

    max_lines = max(1, int((height - padding_y * 2) // (font_size + 1)))
    lines = lines[:max_lines]
    block_height = len(lines) * (font_size + 1)

    alignment = getattr(cell.alignment, "horizontal", None) or "general"
    vertical = getattr(cell.alignment, "vertical", None) or "center"

    if vertical == "top":
        text_y = y + height - padding_y - font_size
    elif vertical == "bottom":
        text_y = y + padding_y + block_height - font_size
    else:
        text_y = y + (height + block_height) / 2 - font_size

    c.setFont(FONT_KR, font_size)
    c.setFillColor(colors.black)

    for idx, line in enumerate(lines):
        if alignment in {"right", "distributed"}:
            c.drawRightString(x + width - padding_x, text_y - idx * (font_size + 1), line)
        elif alignment == "center":
            c.drawCentredString(x + width / 2, text_y - idx * (font_size + 1), line)
        else:
            c.drawString(x + padding_x, text_y - idx * (font_size + 1), line)


def _draw_images(
    c: canvas.Canvas,
    ws,
    x_positions: dict[int, float],
    y_positions: dict[int, float],
    min_col: int,
    min_row: int,
    scale: float,
    x_offset: float,
    y_top: float,
) -> None:
    for image in getattr(ws, "_images", []):
        anchor = image.anchor
        start_col = anchor._from.col + 1
        end_col = anchor.to.col + 1
        start_row = anchor._from.row + 1
        end_row = anchor.to.row + 1
        if end_col < min_col or end_row < min_row:
            continue
        x1 = x_offset + x_positions[max(start_col, min_col) - 1] * scale
        x2 = x_offset + x_positions[end_col] * scale
        y1 = y_top - y_positions[end_row] * scale
        y2 = y_top - y_positions[max(start_row, min_row) - 1] * scale
        width = max(1.0, x2 - x1)
        height = max(1.0, y2 - y1)

        image_ref = getattr(image, "ref", None)
        if image_ref is None:
            continue
        if hasattr(image_ref, "seek"):
            image_ref.seek(0)
            image_reader = ImageReader(image_ref)
        else:
            image_reader = ImageReader(BytesIO(image_ref))
        c.drawImage(image_reader, x1, y1, width=width, height=height, mask="auto")


def export_sheet_to_pdf(workbook_path: Path, sheet_name: str, output_pdf_path: Path) -> Path:
    _register_fonts()
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]

    min_row, min_col, max_row, max_col = _used_bounds(ws)
    x_positions, y_positions, sheet_width, sheet_height = _positions(ws, min_row, min_col, max_row, max_col)

    page_width, page_height = A4
    left_margin = 12 * mm
    right_margin = 12 * mm
    top_margin = 12 * mm
    bottom_margin = 12 * mm
    content_width = page_width - left_margin - right_margin
    content_height = page_height - top_margin - bottom_margin

    scale = min(content_width / sheet_width, content_height / sheet_height)
    x_offset = left_margin + (content_width - sheet_width * scale) / 2
    y_top = page_height - top_margin - (content_height - sheet_height * scale) / 2

    merged_map: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    merged_hidden: set[tuple[int, int]] = set()
    for merged_range in ws.merged_cells.ranges:
        range_min_col, range_min_row, max_col_range, max_row_range = range_boundaries(str(merged_range))
        merged_map[(range_min_row, range_min_col)] = (range_min_row, range_min_col, max_row_range, max_col_range)
        for row_idx in range(range_min_row, max_row_range + 1):
            for col_idx in range(range_min_col, max_col_range + 1):
                if (row_idx, col_idx) != (range_min_row, range_min_col):
                    merged_hidden.add((row_idx, col_idx))

    output_pdf_path.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(output_pdf_path), pagesize=A4)

    for row_idx in range(min_row, max_row + 1):
        for col_idx in range(min_col, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            x = x_offset + x_positions[col_idx - 1] * scale
            y = y_top - y_positions[row_idx] * scale
            width = (x_positions[col_idx] - x_positions[col_idx - 1]) * scale
            height = (y_positions[row_idx] - y_positions[row_idx - 1]) * scale

            fill = _fill_color(cell.fill)
            if fill is not None:
                c.setFillColor(fill)
                c.rect(x, y, width, height, stroke=0, fill=1)

            border = cell.border
            _draw_border_line(c, x, y + height, x + width, y + height, _border_width(border.top.style))
            _draw_border_line(c, x, y, x + width, y, _border_width(border.bottom.style))
            _draw_border_line(c, x, y, x, y + height, _border_width(border.left.style))
            _draw_border_line(c, x + width, y, x + width, y + height, _border_width(border.right.style))

    for row_idx in range(min_row, max_row + 1):
        for col_idx in range(min_col, max_col + 1):
            if (row_idx, col_idx) in merged_hidden:
                continue
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell, MergedCell):
                continue

            merged = merged_map.get((row_idx, col_idx))
            if merged is not None:
                _, _, end_row, end_col = merged
                width = (x_positions[end_col] - x_positions[col_idx - 1]) * scale
                height = (y_positions[end_row] - y_positions[row_idx - 1]) * scale
            else:
                width = (x_positions[col_idx] - x_positions[col_idx - 1]) * scale
                height = (y_positions[row_idx] - y_positions[row_idx - 1]) * scale

            x = x_offset + x_positions[col_idx - 1] * scale
            y = y_top - y_positions[row_idx - 1] * scale - height
            _draw_text(c, cell, x, y, width, height)

    _draw_images(c, ws, x_positions, y_positions, min_col, min_row, scale, x_offset, y_top)
    c.save()
    return output_pdf_path
