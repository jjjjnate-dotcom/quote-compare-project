from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date
import math
from pathlib import Path
import re
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .excel_utils import (
    SOURCE_SHEET_NAME,
    apply_row_merges,
    clear_row_values,
    copy_row_style,
    copy_sheet_content,
    detect_item_count,
)
from .quote_source_metadata import QuoteSourceMetadata, read_metadata_sheet


class QuoteGenerationError(Exception):
    pass


@dataclass
class CompareSheetSpec:
    sheet_name: str
    item_start_row: int
    template_capacity: int
    template_total_row: int
    style_row: int
    max_col: int


@dataclass
class SupplierInfo:
    trade_name: str
    representative: str
    business_number: str
    address: str
    tel: str
    fax: str


INVALID_SHEET_TITLE_CHARS = re.compile(r"[\\/*?:\[\]]")
ITEM_SPEC_RE = re.compile(r"^(?P<name>.+?)\s*(?P<spec>\[[^\]]+\])\s*$")


class QuoteGenerator:
    def __init__(self, template_path: Path):
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise QuoteGenerationError(f"Template file not found: {self.template_path}")

    def generate(
        self,
        source_quote_path: Path,
        output_path: Path,
        company1_name: str,
        company2_name: str,
        company1_rate: float,
        company2_rate: float,
        vat_rate: float,
        include_company3: bool = False,
        company3_name: str | None = None,
        company3_rate: float | None = None,
        company3_supplier: SupplierInfo | None = None,
    ) -> Path:
        try:
            template_wb = load_workbook(self.template_path)
        except (BadZipFile, InvalidFileException) as exc:
            raise QuoteGenerationError(f"Cannot open template file: {self.template_path.name}") from exc

        try:
            source_wb = load_workbook(source_quote_path, data_only=False)
        except (BadZipFile, InvalidFileException) as exc:
            raise QuoteGenerationError(
                "Cannot open uploaded workbook. Check .xlsx/.xlsm format and file integrity."
            ) from exc

        source_ws = source_wb.worksheets[0]
        source_metadata = read_metadata_sheet(source_wb)
        item_count = detect_item_count(source_ws)
        if item_count == 0:
            raise QuoteGenerationError("No items found in the first sheet of uploaded workbook.")

        self._replace_source_sheet(template_wb, source_ws)

        compare_sheets = template_wb.worksheets[1:]
        if len(compare_sheets) < 2:
            raise QuoteGenerationError("Template workbook structure is invalid.")

        raw_titles = [company1_name, company2_name]
        if include_company3:
            if len(compare_sheets) < 3:
                raise QuoteGenerationError("The template does not contain a sheet for company3.")
            if company3_rate is None or company3_supplier is None:
                raise QuoteGenerationError("Company3 information is incomplete.")
            raw_titles.append(company3_name or "업체3")

        sheet_titles = self._make_unique_sheet_titles(raw_titles)

        compare_sheets[0].title = sheet_titles[0]
        compare_sheets[1].title = sheet_titles[1]

        self._fill_geoseong_sheet(
            compare_sheets[0],
            source_ws,
            source_metadata,
            item_count,
            company1_rate,
            vat_rate,
        )
        self._fill_haegwang_sheet(
            compare_sheets[1],
            source_ws,
            source_metadata,
            item_count,
            company2_rate,
            vat_rate,
        )

        if include_company3:
            compare_sheets[2].title = sheet_titles[2]
            self._fill_company3_sheet(
                compare_sheets[2],
                source_ws,
                source_metadata,
                item_count,
                company3_rate,
                vat_rate,
                company3_supplier,
            )
        elif len(compare_sheets) >= 3:
            template_wb.remove(compare_sheets[2])

        output_path.parent.mkdir(parents=True, exist_ok=True)
        template_wb.save(output_path)
        return output_path

    @staticmethod
    def _sanitize_sheet_title(raw_title: str, fallback: str) -> str:
        title = (raw_title or "").strip()
        title = INVALID_SHEET_TITLE_CHARS.sub("_", title)
        title = title.strip("'")
        if not title:
            title = fallback
        return title[:31]

    @classmethod
    def _make_unique_sheet_titles(cls, raw_titles: list[str]) -> list[str]:
        titles: list[str] = []
        used_titles: set[str] = set()

        for index, raw_title in enumerate(raw_titles, start=1):
            base_title = cls._sanitize_sheet_title(raw_title, f"Company{index}")
            candidate = base_title
            suffix_index = 2

            while candidate in used_titles:
                suffix = f" ({suffix_index})"
                candidate = f"{base_title[:31 - len(suffix)]}{suffix}"
                suffix_index += 1

            used_titles.add(candidate)
            titles.append(candidate)

        return titles

    @staticmethod
    def _to_float(value) -> float | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.replace(",", "").strip()
            if cleaned == "":
                return None
            try:
                return float(cleaned)
            except ValueError:
                return None
        return None

    @staticmethod
    def _round_to_hundred_half_up(value: float) -> float:
        factor = 100.0
        if value >= 0:
            return math.floor(value / factor + 0.5) * factor
        return -math.floor(abs(value) / factor + 0.5) * factor

    @staticmethod
    def _split_name_and_spec(raw_name) -> tuple[str, str]:
        name = str(raw_name or "").strip()
        if not name:
            return "", ""

        match = ITEM_SPEC_RE.match(name)
        if not match:
            return name, ""

        return match.group("name").strip(), match.group("spec").strip()

    @staticmethod
    def _guess_unit(item_name: str) -> str:
        compact_name = item_name.replace(" ", "")
        if any(keyword in compact_name for keyword in ("배송", "운반", "설치", "공사")):
            return "건"
        return "EA"

    @staticmethod
    def _format_total_text(amount: float) -> str:
        return f"￦ {amount:,.0f} 원정"

    @staticmethod
    def _format_quote_date(today: date) -> str:
        return f"견적일자: {today.year}년 {today.month}월 {today.day}일"

    def _replace_source_sheet(self, wb: Workbook, source_ws) -> None:
        old_ws = wb[wb.sheetnames[0]]
        wb.remove(old_ws)
        new_ws = wb.create_sheet(title=SOURCE_SHEET_NAME, index=0)
        copy_sheet_content(source_ws, new_ws)

    def _apply_recipient_to_geoseong(self, ws, metadata: QuoteSourceMetadata) -> None:
        lines: list[str] = []
        if metadata.recipient_name:
            lines.append(metadata.recipient_name)

        contact_parts: list[str] = []
        if metadata.recipient_phone:
            contact_parts.append(f"TEL {metadata.recipient_phone}")
        if metadata.recipient_fax:
            contact_parts.append(f"FAX {metadata.recipient_fax}")
        if contact_parts:
            lines.append(" / ".join(contact_parts))

        cell = ws.cell(6, 2)
        cell.value = "\n".join(lines) if lines else ""
        alignment = copy(cell.alignment)
        alignment.wrap_text = True
        cell.alignment = alignment

        current_height = ws.row_dimensions[6].height or 15.0
        if lines:
            ws.row_dimensions[6].height = max(current_height, 30.0 if len(lines) > 1 else 20.0)

    def _apply_recipient_to_haegwang(self, ws, metadata: QuoteSourceMetadata) -> None:
        ws.cell(2, 6).value = metadata.recipient_name or ""
        ws.cell(5, 6).value = metadata.recipient_phone or ""
        ws.cell(6, 6).value = metadata.recipient_fax or ""

    def _apply_supplier_to_company3(self, ws, supplier_info: SupplierInfo) -> None:
        ws.cell(4, 1).value = f"상호: {supplier_info.trade_name}"
        ws.cell(5, 1).value = f"대표: {supplier_info.representative}"
        ws.cell(6, 1).value = f"사업자번호: {supplier_info.business_number}"
        ws.cell(7, 1).value = f"주소: {supplier_info.address}"
        ws.cell(8, 1).value = f"TEL: {supplier_info.tel} / FAX: {supplier_info.fax}"

        for row_idx in (7, 8):
            cell = ws.cell(row_idx, 1)
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            cell.alignment = alignment

    def _apply_recipient_to_company3(self, ws, metadata: QuoteSourceMetadata) -> None:
        ws.cell(4, 6).value = f"상호: {metadata.recipient_name or ''}"
        ws.cell(5, 6).value = "대표:"
        ws.cell(6, 6).value = "사업자번호:"
        ws.cell(7, 6).value = "주소:"

        contact_text = ""
        if metadata.recipient_phone or metadata.recipient_fax:
            contact_text = f"TEL: {metadata.recipient_phone or ''} / FAX: {metadata.recipient_fax or ''}"
        ws.cell(8, 6).value = contact_text

        for row_idx in (7, 8):
            cell = ws.cell(row_idx, 6)
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            cell.alignment = alignment

    def _fill_geoseong_sheet(
        self,
        ws,
        source_ws,
        source_metadata: QuoteSourceMetadata,
        item_count: int,
        rate: float,
        vat_rate: float,
    ) -> None:
        spec = CompareSheetSpec(
            sheet_name=ws.title,
            item_start_row=11,
            template_capacity=23,
            template_total_row=34,
            style_row=11,
            max_col=13,
        )
        extra_rows = max(0, item_count - spec.template_capacity)
        if extra_rows:
            ws.insert_rows(spec.template_total_row, extra_rows)
            for row_idx in range(spec.template_total_row, spec.template_total_row + extra_rows):
                copy_row_style(ws, spec.style_row, row_idx, spec.max_col)
                apply_row_merges(ws, row_idx, [(2, 5), (9, 10), (11, 12)])

        total_row = spec.template_total_row + extra_rows
        last_item_row = spec.item_start_row + item_count - 1
        supply_total = 0.0

        for row_idx in range(spec.item_start_row, last_item_row + 1):
            source_row = row_idx - spec.item_start_row + 2
            source_name = source_ws.cell(source_row, 1).value
            source_qty = source_ws.cell(source_row, 2).value
            source_price = source_ws.cell(source_row, 3).value

            qty_num = self._to_float(source_qty)
            price_num = self._to_float(source_price)
            adjusted_price = self._round_to_hundred_half_up(price_num * (1 + rate)) if price_num is not None else None
            supply_amount = adjusted_price * qty_num if adjusted_price is not None and qty_num is not None else None
            vat_amount = supply_amount * vat_rate if supply_amount is not None else None

            ws.cell(row_idx, 1).value = row_idx - spec.item_start_row + 1
            ws.cell(row_idx, 2).value = source_name
            ws.cell(row_idx, 6).value = source_qty
            ws.cell(row_idx, 8).value = adjusted_price
            ws.cell(row_idx, 9).value = supply_amount
            ws.cell(row_idx, 11).value = vat_amount
            ws.cell(row_idx, 13).value = None

            if supply_amount is not None:
                supply_total += supply_amount

        for row_idx in range(last_item_row + 1, total_row):
            clear_row_values(ws, row_idx, [1, 2, 6, 8, 9, 11, 13])

        ws.cell(total_row, 1).value = "TOTAL"
        ws.cell(total_row, 9).value = supply_total
        ws.cell(total_row, 11).value = supply_total * vat_rate
        ws.cell(9, 10).value = ws.cell(total_row, 9).value + ws.cell(total_row, 11).value
        ws.cell(4, 2).value = date.today()
        ws.cell(9, 6).value = ws.cell(9, 10).value
        self._apply_recipient_to_geoseong(ws, source_metadata)

    def _fill_haegwang_sheet(
        self,
        ws,
        source_ws,
        source_metadata: QuoteSourceMetadata,
        item_count: int,
        rate: float,
        vat_rate: float,
    ) -> None:
        spec = CompareSheetSpec(
            sheet_name=ws.title,
            item_start_row=15,
            template_capacity=20,
            template_total_row=36,
            style_row=15,
            max_col=32,
        )
        extra_rows = max(0, item_count - spec.template_capacity)
        if extra_rows:
            ws.insert_rows(spec.template_total_row, extra_rows)
            for row_idx in range(spec.template_total_row, spec.template_total_row + extra_rows):
                copy_row_style(ws, spec.style_row, row_idx, spec.max_col)
                apply_row_merges(ws, row_idx, [(4, 15), (16, 17), (19, 21), (22, 25), (26, 28), (29, 32)])

        total_row = spec.template_total_row + extra_rows
        last_item_row = spec.item_start_row + item_count - 1
        supply_total = 0.0

        for row_idx in range(spec.item_start_row, last_item_row + 1):
            source_row = row_idx - spec.item_start_row + 2
            source_name = source_ws.cell(source_row, 1).value
            source_qty = source_ws.cell(source_row, 2).value
            source_price = source_ws.cell(source_row, 3).value

            qty_num = self._to_float(source_qty)
            price_num = self._to_float(source_price)
            adjusted_price = self._round_to_hundred_half_up(price_num * (1 + rate)) if price_num is not None else None
            supply_amount = adjusted_price * qty_num if adjusted_price is not None and qty_num is not None else None
            vat_amount = supply_amount * vat_rate if supply_amount is not None else None

            ws.cell(row_idx, 3).value = row_idx - spec.item_start_row + 1
            ws.cell(row_idx, 4).value = source_name
            ws.cell(row_idx, 18).value = source_qty
            ws.cell(row_idx, 19).value = adjusted_price
            ws.cell(row_idx, 22).value = supply_amount
            ws.cell(row_idx, 26).value = vat_amount
            ws.cell(row_idx, 29).value = None

            if supply_amount is not None:
                supply_total += supply_amount

        for row_idx in range(last_item_row + 1, total_row):
            clear_row_values(ws, row_idx, [3, 4, 18, 19, 22, 26, 29])

        ws.cell(total_row, 3).value = "TOTAL"
        ws.cell(total_row, 5).value = supply_total
        ws.cell(total_row, 10).value = "VAT"
        ws.cell(total_row, 14).value = supply_total * vat_rate
        ws.cell(total_row, 19).value = "SUM(Total)"
        ws.cell(total_row, 25).value = ws.cell(total_row, 5).value + ws.cell(total_row, 14).value
        ws.cell(8, 6).value = date.today()
        self._apply_recipient_to_haegwang(ws, source_metadata)

    def _fill_company3_sheet(
        self,
        ws,
        source_ws,
        source_metadata: QuoteSourceMetadata,
        item_count: int,
        rate: float,
        vat_rate: float,
        supplier_info: SupplierInfo,
    ) -> None:
        spec = CompareSheetSpec(
            sheet_name=ws.title,
            item_start_row=11,
            template_capacity=20,
            template_total_row=31,
            style_row=11,
            max_col=10,
        )

        extra_rows = max(0, item_count - spec.template_capacity)
        if extra_rows:
            ws.insert_rows(spec.template_total_row, extra_rows)
            for row_idx in range(spec.template_total_row, spec.template_total_row + extra_rows):
                copy_row_style(ws, spec.style_row, row_idx, spec.max_col)

        total_row = spec.template_total_row + extra_rows
        last_item_row = spec.item_start_row + item_count - 1
        supply_total = 0.0

        for row_idx in range(spec.item_start_row, last_item_row + 1):
            source_row = row_idx - spec.item_start_row + 2
            raw_name = source_ws.cell(source_row, 1).value
            source_qty = source_ws.cell(source_row, 2).value
            source_price = source_ws.cell(source_row, 3).value

            item_name, item_spec = self._split_name_and_spec(raw_name)
            qty_num = self._to_float(source_qty)
            price_num = self._to_float(source_price)
            adjusted_price = self._round_to_hundred_half_up(price_num * (1 + rate)) if price_num is not None else None
            supply_amount = adjusted_price * qty_num if adjusted_price is not None and qty_num is not None else None
            vat_amount = supply_amount * vat_rate if supply_amount is not None else None
            total_amount = supply_amount + vat_amount if supply_amount is not None and vat_amount is not None else None

            ws.cell(row_idx, 1).value = row_idx - spec.item_start_row + 1
            ws.cell(row_idx, 2).value = item_name
            ws.cell(row_idx, 3).value = item_spec
            ws.cell(row_idx, 4).value = self._guess_unit(item_name)
            ws.cell(row_idx, 5).value = source_qty
            ws.cell(row_idx, 6).value = adjusted_price
            ws.cell(row_idx, 7).value = supply_amount
            ws.cell(row_idx, 8).value = vat_amount
            ws.cell(row_idx, 9).value = total_amount
            ws.cell(row_idx, 10).value = None

            if supply_amount is not None:
                supply_total += supply_amount

        for row_idx in range(last_item_row + 1, total_row):
            clear_row_values(ws, row_idx, list(range(1, 11)))

        vat_total = supply_total * vat_rate
        grand_total = supply_total + vat_total
        today = date.today()

        ws.cell(2, 1).value = f"견적번호: {today:%Y%m%d}"
        ws.cell(2, 6).value = self._format_quote_date(today)
        ws.cell(9, 4).value = self._format_total_text(grand_total)
        ws.cell(total_row, 1).value = "소  계"
        ws.cell(total_row, 7).value = supply_total
        ws.cell(total_row, 8).value = vat_total
        ws.cell(total_row, 9).value = grand_total

        self._apply_supplier_to_company3(ws, supplier_info)
        self._apply_recipient_to_company3(ws, source_metadata)
