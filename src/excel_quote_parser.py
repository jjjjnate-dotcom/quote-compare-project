from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook

from .quote_source_metadata import QuoteSourceMetadata, write_metadata_sheet


class ExcelQuoteParseError(Exception):
    pass


@dataclass
class QuoteItem:
    name: str
    qty: float
    unit_price: float


PHONE_RE = re.compile(r"\d{2,4}-\d{3,4}-\d{4}")


def _normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip().replace(" ", "")


def _to_number(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = re.sub(r"[^0-9.\-]", "", text.replace(",", ""))
    if text in {"", "-", ".", "-."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _is_total_row(name_value) -> bool:
    text = _normalize_text(name_value)
    return any(keyword in text for keyword in ("총합계", "합계", "계"))


def _find_header_columns(ws) -> tuple[int, int, int, int] | None:
    max_row = min(ws.max_row, 80)
    max_col = min(ws.max_column, 30)

    for row_idx in range(1, max_row + 1):
        name_col = qty_col = price_col = None
        for col_idx in range(1, max_col + 1):
            text = _normalize_text(ws.cell(row_idx, col_idx).value)
            if not text:
                continue
            if name_col is None and ("품목명" in text or "품명" in text):
                name_col = col_idx
            if qty_col is None and "수량" in text:
                qty_col = col_idx
            if price_col is None and "단가" in text:
                price_col = col_idx

        if name_col and qty_col and price_col:
            return row_idx, name_col, qty_col, price_col

    return None


def _extract_from_header_table(ws) -> list[QuoteItem]:
    header = _find_header_columns(ws)
    if not header:
        return []

    header_row, name_col, qty_col, price_col = header
    items: list[QuoteItem] = []
    blank_streak = 0

    for row_idx in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row_idx, name_col).value
        qty = ws.cell(row_idx, qty_col).value
        unit_price = ws.cell(row_idx, price_col).value

        if _is_total_row(name):
            break

        if name in (None, "") and qty in (None, "") and unit_price in (None, ""):
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue

        blank_streak = 0
        qty_num = _to_number(qty)
        price_num = _to_number(unit_price)
        name_text = str(name).strip() if name is not None else ""

        if not name_text or qty_num is None or price_num is None:
            continue
        if qty_num <= 0 or price_num <= 0:
            continue

        items.append(QuoteItem(name=name_text, qty=qty_num, unit_price=price_num))

    return items


def _extract_from_simple_columns(ws) -> list[QuoteItem]:
    items: list[QuoteItem] = []
    blank_streak = 0

    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row_idx, 1).value
        qty = ws.cell(row_idx, 2).value
        unit_price = ws.cell(row_idx, 3).value

        if name in (None, "") and qty in (None, "") and unit_price in (None, ""):
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue

        blank_streak = 0
        if _is_total_row(name):
            break

        qty_num = _to_number(qty)
        price_num = _to_number(unit_price)
        name_text = str(name).strip() if name is not None else ""

        if not name_text or qty_num is None or price_num is None:
            continue
        if qty_num <= 0 or price_num <= 0:
            continue

        items.append(QuoteItem(name=name_text, qty=qty_num, unit_price=price_num))

    return items


def _extract_inline_recipient(text: str) -> str | None:
    match = re.search(r"수신(?:\s*-\s*거래처)?\s*[:：]\s*(.+)", text)
    if match:
        return match.group(1).strip()
    return None


def _extract_inline_phone_pair(text: str) -> tuple[str | None, str | None]:
    match = re.search(
        r"TEL\s*/\s*FAX\s*[:：]?\s*(\d{2,4}-\d{3,4}-\d{4})\s*/\s*(\d{2,4}-\d{3,4}-\d{4})",
        text,
        re.IGNORECASE,
    )
    if match:
        return match.group(1), match.group(2)
    return None, None


def _neighbor_text(ws, row_idx: int, col_idx: int) -> str | None:
    for next_col in range(col_idx + 1, min(ws.max_column, col_idx + 3) + 1):
        value = ws.cell(row_idx, next_col).value
        if value not in (None, ""):
            return str(value).strip()
    return None


def extract_metadata_from_excel(source_path: Path) -> QuoteSourceMetadata:
    wb = load_workbook(source_path, data_only=True)
    ws = wb.worksheets[0]

    metadata = QuoteSourceMetadata()
    max_row = min(ws.max_row, 40)
    max_col = min(ws.max_column, 12)

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            value = ws.cell(row_idx, col_idx).value
            if value in (None, ""):
                continue

            text = str(value).strip()
            normalized = _normalize_text(text)

            if metadata.recipient_name is None:
                recipient = _extract_inline_recipient(text)
                if recipient:
                    metadata.recipient_name = recipient
                elif normalized in {"수신", "수신거래처"}:
                    metadata.recipient_name = _neighbor_text(ws, row_idx, col_idx)

            inline_phone, inline_fax = _extract_inline_phone_pair(text)
            if inline_phone and metadata.recipient_phone is None:
                metadata.recipient_phone = inline_phone
            if inline_fax and metadata.recipient_fax is None:
                metadata.recipient_fax = inline_fax

            if metadata.recipient_phone is None and normalized in {"전화", "전화번호", "tel"}:
                neighbor = _neighbor_text(ws, row_idx, col_idx)
                if neighbor:
                    match = PHONE_RE.search(neighbor)
                    if match:
                        metadata.recipient_phone = match.group(0)

            if metadata.recipient_fax is None and normalized in {"팩스", "fax"}:
                neighbor = _neighbor_text(ws, row_idx, col_idx)
                if neighbor:
                    match = PHONE_RE.search(neighbor)
                    if match:
                        metadata.recipient_fax = match.group(0)

    return metadata


def extract_items_from_excel(source_path: Path) -> list[QuoteItem]:
    wb = load_workbook(source_path, data_only=True)
    ws = wb.worksheets[0]

    items = _extract_from_header_table(ws)
    if not items:
        items = _extract_from_simple_columns(ws)

    if not items:
        raise ExcelQuoteParseError(
            "엑셀에서 품목/수량/단가를 찾지 못했습니다. 표 헤더(품목명, 수량, 단가)를 확인해 주세요."
        )

    return items


def convert_excel_to_source_workbook(source_path: Path, output_xlsx_path: Path) -> Path:
    items = extract_items_from_excel(source_path)
    metadata = extract_metadata_from_excel(source_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "본견적"
    ws.cell(1, 1).value = "품목명"
    ws.cell(1, 2).value = "수량"
    ws.cell(1, 3).value = "단가"

    for row_idx, item in enumerate(items, start=2):
        ws.cell(row_idx, 1).value = item.name
        ws.cell(row_idx, 2).value = item.qty
        ws.cell(row_idx, 3).value = item.unit_price

    write_metadata_sheet(wb, metadata)
    output_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx_path)
    return output_xlsx_path
