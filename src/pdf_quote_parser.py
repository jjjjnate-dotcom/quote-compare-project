from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re

from openpyxl import Workbook
from pypdf import PdfReader

from .quote_source_metadata import QuoteSourceMetadata, write_metadata_sheet


class PdfQuoteParseError(Exception):
    pass


@dataclass
class QuoteItem:
    name: str
    qty: float
    unit_price: float


ITEM_LINE_WITH_INDEX_RE = re.compile(
    r"^\s*(?P<idx>\d+)\s+"
    r"(?P<name>.+?)\s+"
    r"(?P<qty>\d[\d,]*(?:\.\d+)?)\s+"
    r"(?P<unit>\d[\d,]*(?:\.\d+)?)\s+"
    r"(?P<supply>\d[\d,]*(?:\.\d+)?)"
    r"(?:\s+\d[\d,]*(?:\.\d+)?)?\s*$"
)

ITEM_LINE_NO_INDEX_RE = re.compile(
    r"^\s*"
    r"(?P<name>.+?)\s+"
    r"(?P<qty>\d[\d,]*(?:\.\d+)?)\s+"
    r"(?P<unit>\d[\d,]*(?:\.\d+)?)\s+"
    r"(?P<supply>\d[\d,]*(?:\.\d+)?)"
    r"(?:\s+\d[\d,]*(?:\.\d+)?)?\s*$"
)

TOTAL_SUPPLY_RE = re.compile(r"공급가\s*([0-9][0-9,]*)")
PHONE_RE = re.compile(r"\d{2,4}-\d{3,4}-\d{4}")


def _to_number(value: str) -> float:
    return float(value.replace(",", "").strip())


def _normalize_line(line: str) -> str:
    line = line.replace("\u00a0", " ")
    return re.sub(r"\s+", " ", line).strip()


def _is_non_item_line(line: str) -> bool:
    lowered = line.lower()
    blocked_keywords = (
        "합계",
        "총합계",
        "공급가",
        "부가세",
        "견적",
        "순번",
        "품목명",
        "수량",
        "단가",
        "공급가액",
    )
    return any(keyword in line for keyword in blocked_keywords) or lowered.startswith("tel") or lowered.startswith("fax")


def extract_metadata_from_pdf(pdf_path: Path) -> QuoteSourceMetadata:
    reader = PdfReader(str(pdf_path))
    text = "\n".join((page.extract_text() or "") for page in reader.pages)
    lines = [_normalize_line(line) for line in text.splitlines()]
    lines = [line for line in lines if line]

    metadata = QuoteSourceMetadata()

    recipient_match = re.search(r"수\s*신(?:\s*[-:]\s*|\s+)(.+)", text)
    if recipient_match:
        metadata.recipient_name = _normalize_line(recipient_match.group(1)).rstrip("?")

    tel_fax_match = re.search(
        r"TEL\s*/\s*FAX\s*[:：]?\s*(\d{2,4}-\d{3,4}-\d{4})\s*/\s*(\d{2,4}-\d{3,4}-\d{4})",
        text,
        re.IGNORECASE,
    )
    if tel_fax_match:
        metadata.recipient_phone = tel_fax_match.group(1)
        metadata.recipient_fax = tel_fax_match.group(2)

    if metadata.recipient_name is None:
        for line in lines:
            match = re.search(r"수\s*신(?:\s*[-:]\s*|\s+)(.+)", line)
            if match:
                metadata.recipient_name = _normalize_line(match.group(1)).rstrip("?")
                break

    if metadata.recipient_phone is None or metadata.recipient_fax is None:
        for line in lines:
            upper_line = line.upper()
            if "TEL/FAX" in upper_line:
                phones = PHONE_RE.findall(line)
                if len(phones) >= 2:
                    metadata.recipient_phone = metadata.recipient_phone or phones[0]
                    metadata.recipient_fax = metadata.recipient_fax or phones[1]
                    break

    return metadata


def extract_items_from_pdf(pdf_path: Path) -> list[QuoteItem]:
    reader = PdfReader(str(pdf_path))
    text = "\n".join((page.extract_text() or "") for page in reader.pages)
    if not text.strip():
        raise PdfQuoteParseError("PDF에서 텍스트를 읽지 못했습니다. 스캔 PDF일 수 있습니다.")

    lines = [_normalize_line(line) for line in text.splitlines()]
    lines = [line for line in lines if line]

    items: list[QuoteItem] = []
    for line in lines:
        if _is_non_item_line(line):
            continue

        match = ITEM_LINE_WITH_INDEX_RE.match(line) or ITEM_LINE_NO_INDEX_RE.match(line)
        if not match:
            continue

        name = match.group("name").strip()
        qty = _to_number(match.group("qty"))
        unit = _to_number(match.group("unit"))

        if not name or qty <= 0 or unit <= 0:
            continue

        items.append(QuoteItem(name=name, qty=qty, unit_price=unit))

    if not items:
        raise PdfQuoteParseError(
            "PDF에서 품목/수량/단가를 찾지 못했습니다. 표 형태 또는 텍스트 인식 상태를 확인해 주세요."
        )

    expected_supply = None
    total_match = TOTAL_SUPPLY_RE.search(text)
    if total_match:
        expected_supply = _to_number(total_match.group(1))

    if expected_supply is not None:
        parsed_supply = sum(item.qty * item.unit_price for item in items)
        tolerance = max(1000.0, expected_supply * 0.03)
        if abs(parsed_supply - expected_supply) > tolerance:
            raise PdfQuoteParseError(
                "PDF 파싱 검증에 실패했습니다. 품목 합계가 문서 공급가와 크게 다릅니다."
            )

    return items


def convert_pdf_to_source_workbook(pdf_path: Path, output_xlsx_path: Path) -> Path:
    items = extract_items_from_pdf(pdf_path)
    metadata = extract_metadata_from_pdf(pdf_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "본견적"
    ws.cell(1, 1).value = "품목명"
    ws.cell(1, 2).value = "수량"
    ws.cell(1, 3).value = "단가"

    for row_idx, item in enumerate(items, start=2):
        ws.cell(row_idx, 1).value = item.name
        ws.cell(row_idx, 2).value = item.qty
        ws.cell(row_idx, 3).value = item.unit_price

    write_metadata_sheet(wb, metadata)
    output_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx_path)
    return output_xlsx_path
