from __future__ import annotations

from copy import copy
from typing import Iterable

from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

SOURCE_SHEET_NAME = "자재피아기본견적"


def is_effectively_blank(value) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def copy_sheet_content(src_ws, dst_ws) -> None:
    dst_ws.sheet_format = copy(src_ws.sheet_format)
    dst_ws.sheet_properties = copy(src_ws.sheet_properties)
    dst_ws.page_margins = copy(src_ws.page_margins)
    dst_ws.page_setup = copy(src_ws.page_setup)
    dst_ws.print_options = copy(src_ws.print_options)
    dst_ws.freeze_panes = src_ws.freeze_panes

    try:
        dst_ws.print_area = src_ws.print_area
    except Exception:
        pass

    for key, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[key] = copy(dim)

    for key, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[key] = copy(dim)

    for row in src_ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            new_cell = dst_ws[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell._style = copy(cell._style)
            new_cell.number_format = cell.number_format
            new_cell.font = copy(cell.font)
            new_cell.fill = copy(cell.fill)
            new_cell.border = copy(cell.border)
            new_cell.alignment = copy(cell.alignment)
            new_cell.protection = copy(cell.protection)
            if cell.hyperlink:
                new_cell._hyperlink = copy(cell.hyperlink)
            if cell.comment:
                new_cell.comment = copy(cell.comment)

    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))


def detect_item_count(ws, start_row: int = 2, name_col: int = 1, qty_col: int = 2, price_col: int = 3) -> int:
    row = start_row
    count = 0
    while True:
        values = [ws.cell(row, c).value for c in (name_col, qty_col, price_col)]
        if all(is_effectively_blank(v) for v in values):
            return count
        count += 1
        row += 1


def copy_row_style(ws, src_row: int, target_row: int, max_col: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        src_cell = ws.cell(src_row, col)
        dst_cell = ws.cell(target_row, col)
        if src_cell.has_style:
            dst_cell._style = copy(src_cell._style)
        dst_cell.number_format = src_cell.number_format
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.protection = copy(src_cell.protection)


def clear_row_values(ws, row: int, columns: Iterable[int]) -> None:
    for col in columns:
        ws.cell(row, col).value = None


def apply_row_merges(ws, row: int, merge_specs: list[tuple[int, int]]) -> None:
    for start_col, end_col in merge_specs:
        ws.merge_cells(f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}")
